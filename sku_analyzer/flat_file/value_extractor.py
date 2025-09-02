"""Value extraction module for XLSM template data.

This module provides focused functionality for extracting valid values
from template columns with high-performance optimization.
"""

from __future__ import annotations

import logging
import re
import time
from pathlib import Path
from typing import Dict, List, Tuple, Union

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from .data_structures import ExtractionMetrics, FieldValidation


class ValueExtractor:
    """High-performance value extractor for template columns."""
    
    # Dynamic column detection will replace this
    VALUE_COLUMNS = ['C', 'D', 'E', 'F']  # Minimum range, expanded dynamically
    
    def __init__(self, enable_performance_monitoring: bool = True) -> None:
        """Initialize value extractor.
        
        Args:
            enable_performance_monitoring: Whether to track performance metrics
        """
        self.logger = logging.getLogger(__name__)
        self.enable_monitoring = enable_performance_monitoring
        
        # Precompiled patterns for performance
        self.unicode_pattern = re.compile(r'[^\x00-\x7F]')
    
    def _detect_data_column_range(self, worksheet: Worksheet, sample_rows: List[int]) -> Tuple[int, int]:
        """Dynamically detect the actual data column range by scanning for non-empty cells.
        
        Args:
            worksheet: Worksheet to analyze
            sample_rows: List of row indices to sample for column detection
            
        Returns:
            Tuple of (min_col_index, max_col_index) 1-based
        """
        min_col = 3  # Start from column C
        max_col = 3
        
        # Use worksheet's actual max column or reasonable limit
        max_scan_col = min(worksheet.max_column, 1000)
        
        # Sample a few rows to determine the actual data range
        for row_idx in sample_rows[:10]:  # Sample first 10 rows
            col_num = 3  # Start from C
            consecutive_empty = 0
            
            while col_num <= max_scan_col:
                col_letter = get_column_letter(col_num)
                cell_value = worksheet[f'{col_letter}{row_idx}'].value
                
                if cell_value and str(cell_value).strip():
                    max_col = max(max_col, col_num)
                    consecutive_empty = 0
                    col_num += 1
                else:
                    consecutive_empty += 1
                    # Stop if we find 10 consecutive empty columns
                    if consecutive_empty >= 10:
                        break
                    col_num += 1
        
        return min_col, max_col
    
    def extract_values_from_valid_values_sheet(
        self, 
        worksheet: Worksheet
    ) -> Dict[str, FieldValidation]:
        """Extract values from 'Gültige Werte' worksheet with dynamic column detection.
        
        Args:
            worksheet: 'Gültige Werte' worksheet to extract from
            
        Returns:
            Dictionary mapping field names to validation results
        """
        field_validations = {}
        
        # Use worksheet's actual max column or reasonable limit
        max_scan_col = min(worksheet.max_column, 1000)
        
        # Scan worksheet for field rows
        for row in range(2, min(worksheet.max_row + 1, 200)):
            field_name_cell = worksheet[f'B{row}'].value
            if not field_name_cell:
                continue
                
            field_name_str = str(field_name_cell).strip()
            if not field_name_str:
                continue
            
            # Extract base field name (remove "- [ pants ]" suffix)
            base_field_name = field_name_str.split(' - [')[0].strip()
            
            # Dynamic column detection for this row
            values_found = []
            col_num = 3  # Start from C
            consecutive_empty = 0
            
            while col_num <= max_scan_col:
                col_letter = get_column_letter(col_num)
                cell_value = worksheet[f'{col_letter}{row}'].value
                
                if cell_value and str(cell_value).strip():
                    values_found.append(str(cell_value).strip())
                    consecutive_empty = 0
                    col_num += 1
                else:
                    consecutive_empty += 1
                    # Stop if we find 10 consecutive empty columns (allows for sparse data)
                    if consecutive_empty >= 10:
                        break
                    col_num += 1
            
            # Create field validation if values found
            if values_found:
                values_with_flags = [(val, {'is_empty': False, 'is_unicode': False, 'is_formula': False, 'is_merged': False}) for val in values_found]
                field_validations[base_field_name] = self._create_field_validation(
                    base_field_name, values_with_flags
                )
        
        return field_validations

    def extract_bulk_values(
        self, 
        worksheet: Worksheet, 
        field_mappings: List[Dict[str, any]]
    ) -> Dict[str, FieldValidation]:
        """Extract values using bulk range operations for maximum performance.
        
        Args:
            worksheet: Worksheet to extract from
            field_mappings: List of field mapping dictionaries
            
        Returns:
            Dictionary mapping field names to validation results
        """
        # Build row index to field mapping
        row_to_fields = {}
        for mapping in field_mappings:
            row_idx = mapping['row_index']
            if row_idx not in row_to_fields:
                row_to_fields[row_idx] = []
            row_to_fields[row_idx].append(mapping)
        
        unique_rows = sorted(row_to_fields.keys())
        if not unique_rows:
            return {}
        
        # Dynamically detect actual data column range
        min_row = min(unique_rows)
        max_row = max(unique_rows)
        min_col, max_col = self._detect_data_column_range(worksheet, unique_rows)
        
        # Extract all values within detected range
        bulk_values = {}
        for row_offset, row in enumerate(worksheet.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col,
            values_only=True
        )):
            row_index = min_row + row_offset
            if row_index in row_to_fields:
                # Store all column values dynamically
                row_data = {}
                for col_offset, cell_value in enumerate(row):
                    col_letter = get_column_letter(min_col + col_offset)
                    row_data[col_letter] = cell_value
                bulk_values[row_index] = row_data
        
        # Process all fields using bulk extracted data
        field_validations = {}
        
        for row_idx, mappings in row_to_fields.items():
            row_values = bulk_values.get(row_idx, {})
            
            for mapping in mappings:
                field_name = mapping['technical_name']
                requirement_status = mapping.get('requirement_status', 'optional')
                
                # Extract values from all detected columns
                values_with_flags = []
                for col_letter, raw_value in row_values.items():
                    if raw_value is None or raw_value == '':
                        continue
                    
                    str_value = str(raw_value).strip()
                    if not str_value:
                        continue
                    
                    # Fast classification
                    flags = {
                        'is_empty': False,
                        'is_unicode': ord(max(str_value)) > 127 if str_value else False,
                        'is_formula': str_value.startswith('='),
                        'is_merged': False
                    }
                    
                    values_with_flags.append((str_value, flags))
                
                # Create field validation
                field_validations[field_name] = self._create_field_validation(
                    field_name, values_with_flags
                )
        
        return field_validations
    
    def _create_field_validation(
        self, 
        field_name: str, 
        values_with_flags: List[Tuple[str, Dict[str, bool]]]
    ) -> FieldValidation:
        """Create field validation with optimized processing.
        
        Args:
            field_name: Name of the field
            values_with_flags: List of (value, flags) tuples
            
        Returns:
            FieldValidation object with processing results
        """
        validation = FieldValidation(field_name, 'not_serialized')
        
        # Use sets for deduplication
        valid_values_set = set()
        invalid_values_set = set()
        
        for value, flags in values_with_flags:
            validation.total_values += 1
            
            # Update counters
            if flags['is_unicode']:
                validation.unicode_values += 1
            if flags['is_formula']:
                validation.formula_values += 1
            if flags['is_merged']:
                validation.merged_cell_values += 1
            
            # Validation logic - capture all non-empty values
            if value and value.strip() and not flags.get('is_empty', False):
                valid_values_set.add(value)
        
        # Convert sets to lists
        validation.valid_values = list(valid_values_set)
        validation.invalid_values = list(invalid_values_set)
        
        return validation
    
    def create_extraction_metrics(
        self, 
        operation_name: str,
        duration_ms: float,
        peak_memory_mb: float,
        values_extracted: int,
        columns_processed: int
    ) -> ExtractionMetrics:
        """Create extraction performance metrics.
        
        Args:
            operation_name: Name of the extraction operation
            duration_ms: Operation duration in milliseconds
            peak_memory_mb: Peak memory usage in MB
            values_extracted: Number of values extracted
            columns_processed: Number of columns processed
            
        Returns:
            ExtractionMetrics object with calculated derived metrics
        """
        return ExtractionMetrics(
            operation_name=operation_name,
            duration_ms=duration_ms,
            peak_memory_mb=peak_memory_mb,
            values_extracted=values_extracted,
            columns_processed=columns_processed
        )