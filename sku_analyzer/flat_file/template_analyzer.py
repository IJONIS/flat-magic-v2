"""Modern XLSM template analyzer with clean modular architecture.

Refactored to use focused, single-responsibility modules following
modern Python patterns with full type hints and clean separation of concerns.
"""

from __future__ import annotations

import asyncio
import json
import logging
import time
from pathlib import Path
from typing import Any, Dict, Optional, Union

from openpyxl import load_workbook

from .column_extractor import ColumnExtractor
from .data_structures import TemplateAnalysisResult, TemplateDetectionError
from .header_detector import HeaderDetector
from .performance_monitor import PerformanceMonitor
from .validation_utils import ValidationUtils
from .worksheet_detector import WorksheetDetector


class OptimizedXlsmTemplateAnalyzer:
    """Modern template analyzer using clean modular architecture.
    
    This analyzer coordinates specialized components to perform template analysis
    with clean separation of concerns and modern Python patterns.
    """
    
    def __init__(self, enable_performance_monitoring: bool = True) -> None:
        """Initialize template analyzer with modular components.
        
        Args:
            enable_performance_monitoring: Whether to enable performance tracking
        """
        self.logger = self._setup_logging()
        self.enable_monitoring = enable_performance_monitoring
        
        # Initialize modular components
        self.worksheet_detector = WorksheetDetector()
        self.header_detector = HeaderDetector()
        self.column_extractor = ColumnExtractor(enable_performance_monitoring)
        self.performance_monitor = PerformanceMonitor(enable_performance_monitoring)
        self.validation_utils = ValidationUtils()
    
    def _get_next_job_number(self) -> int:
        """Get the next consecutive job number.
        
        Returns:
            Next available job number
        """
        production_dir = Path("production_output")
        if not production_dir.exists():
            return 1
        
        job_numbers = [
            int(job_dir.name) for job_dir in production_dir.iterdir()
            if job_dir.is_dir() and job_dir.name.isdigit()
        ]
        
        return max(job_numbers, default=0) + 1
    
    def get_latest_job_number(self) -> Optional[int]:
        """Get the most recent job number.
        
        Returns:
            Latest job number or None if no jobs exist
        """
        production_dir = Path("production_output")
        if not production_dir.exists():
            return None
        
        job_numbers = [
            int(job_dir.name) for job_dir in production_dir.iterdir()
            if job_dir.is_dir() and job_dir.name.isdigit()
        ]
        
        return max(job_numbers) if job_numbers else None
    
    def _setup_logging(self) -> logging.Logger:
        """Configure structured logging.
        
        Returns:
            Configured logger instance
        """
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)
        
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - [%(funcName)s] - %(message)s'
            )
            handler.setFormatter(formatter)
            logger.addHandler(handler)
        
        return logger
    
    async def _load_workbook_optimized(self, xlsm_path: Path) -> Any:
        """Load XLSM workbook with optimized settings.
        
        Args:
            xlsm_path: Path to XLSM file
            
        Returns:
            Loaded openpyxl workbook
        """
        loop = asyncio.get_event_loop()
        
        def _load_workbook() -> Any:
            return load_workbook(
                xlsm_path,
                read_only=True,
                keep_vba=False,
                data_only=True,
                keep_links=False
            )
        
        return await loop.run_in_executor(None, _load_workbook)
    
    async def analyze_template(self, xlsm_path: Path, job_id: Optional[Union[int, str]] = None) -> Dict[str, Any]:
        """Analyze XLSM template with modular architecture.
        
        Args:
            xlsm_path: Path to XLSM file
            job_id: Job identifier (creates new if None)
            
        Returns:
            Dictionary containing analysis results and performance metrics
            
        Raises:
            TemplateDetectionError: When template detection fails
            FileNotFoundError: When XLSM file doesn't exist
            ValueError: When invalid parameters provided
        """
        if not xlsm_path.exists():
            raise FileNotFoundError(f"XLSM file not found: {xlsm_path}")
        
        if not xlsm_path.suffix.lower() in ['.xlsm', '.xlsx']:
            raise ValueError(f"Invalid file type: {xlsm_path.suffix}. Expected .xlsm or .xlsx")

        # Use existing job ID or create new one
        if job_id is None:
            job_number = self._get_next_job_number()
        else:
            job_number = int(job_id)

        self.logger.info(f"🔍 Starting modular template analysis for job {job_number}: {xlsm_path}")
        
        performance_tracker = {}
        workbook = None
        
        try:
            # Step 1: Load workbook
            with self.performance_monitor.measure_performance("workbook_loading") as monitor:
                workbook = await self._load_workbook_optimized(xlsm_path)
            if monitor and monitor['metrics']:
                performance_tracker["workbook_loading"] = monitor['metrics']
            
            # Step 2: Detect target worksheet
            with self.performance_monitor.measure_performance("worksheet_detection") as monitor:
                worksheet = self.worksheet_detector.detect_target_worksheet(workbook)
            if monitor and monitor['metrics']:
                performance_tracker["worksheet_detection"] = monitor['metrics']
            
            # Step 3: Detect headers
            with self.performance_monitor.measure_performance("header_detection", worksheet.max_row) as monitor:
                header_row, tech_col, display_col, req_col = self.header_detector.detect_header_row(worksheet)
            if monitor and monitor['metrics']:
                performance_tracker["header_detection"] = monitor['metrics']
            
            # Step 4: Extract column mappings
            with self.performance_monitor.measure_performance("column_extraction", worksheet.max_row) as monitor:
                mappings = self.column_extractor.extract_column_mappings(
                    worksheet, header_row, tech_col, display_col, req_col
                )
            if monitor and monitor['metrics']:
                performance_tracker["column_extraction"] = monitor['metrics']
                
            # Step 5: Generate requirement statistics
            requirement_stats = self.column_extractor.generate_requirement_statistics(mappings)
            
            # Step 6: Create and validate result
            result = TemplateAnalysisResult(
                job_id=job_number,
                worksheet_name=worksheet.title,
                header_row=header_row,
                technical_column=tech_col,
                display_column=display_col,
                column_mappings=mappings,
                total_mappings=len(mappings),
                performance_metrics=performance_tracker,
                requirement_column=req_col,
                requirement_statistics=requirement_stats
            )
            
            self.validation_utils.validate_analysis_result(result)
            
            # Step 7: Save results
            with self.performance_monitor.measure_performance("output_saving") as monitor:
                output_path = await self._save_analysis_result_optimized(result)
            if monitor and monitor['metrics']:
                performance_tracker["output_saving"] = monitor['metrics']
            
            # Step 8: Extract template values (Step 2)
            step2_output_path = None
            try:
                from .template_data_extractor import HighPerformanceTemplateDataExtractor
                
                self.logger.info(f"🔍 Starting Step 2 value extraction for job {job_number}")
                extractor = HighPerformanceTemplateDataExtractor(enable_performance_monitoring=True)
                
                start_time = time.perf_counter()
                step2_output_path = extractor.extract_and_save_values(xlsm_path, job_number)
                step2_duration = (time.perf_counter() - start_time) * 1000
                
                # Record Step 2 performance
                performance_tracker["step2_value_extraction"] = {
                    'duration_ms': step2_duration,
                    'peak_memory_mb': 0,
                    'throughput_rows_per_second': 0
                }
                    
                self.logger.info(f"✅ Step 2 value extraction completed for job {job_number}")
                
            except Exception as e:
                self.logger.error(f"Step 2 value extraction failed for job {job_number}: {e}")
                # Don't fail the entire analysis if Step 2 fails
                pass
            
            # Log performance summary
            total_duration = sum(
                m.duration_ms if hasattr(m, 'duration_ms') else m.get('duration_ms', 0) 
                for m in performance_tracker.values()
            )
            peak_memory = max(
                m.peak_memory_mb if hasattr(m, 'peak_memory_mb') else m.get('peak_memory_mb', 0) 
                for m in performance_tracker.values()
            ) if performance_tracker else 0
            
            self.logger.info(
                f"✅ Template analysis completed for job {job_number}: "
                f"{len(mappings)} mappings in {total_duration:.1f}ms, "
                f"peak memory: {peak_memory:.1f}MB"
            )
            
            analysis_dict = result.to_dict()
            analysis_dict['output_path'] = str(output_path)
            if step2_output_path:
                analysis_dict['step2_output_path'] = str(step2_output_path)
            return analysis_dict
            
        except Exception as e:
            self.logger.error(f"Template analysis failed for job {job_number}: {e}")
            raise TemplateDetectionError(f"Template analysis failed: {e}") from e
        finally:
            # Ensure workbook is closed
            if workbook:
                try:
                    workbook.close()
                except Exception as e:
                    self.logger.warning(f"Error closing workbook: {e}")
    
    async def _save_analysis_result_optimized(self, result: TemplateAnalysisResult) -> Path:
        """Save analysis result with optimized JSON serialization.
        
        Args:
            result: Analysis result to save
            
        Returns:
            Path to saved JSON file
        """
        output_dir = Path("production_output") / str(result.job_id) / "flat_file_analysis"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        output_file = output_dir / "step1_template_columns.json"
        
        try:
            result_dict = result.to_dict()
            
            # Use fastest available JSON library
            try:
                import orjson
                json_bytes = orjson.dumps(
                    result_dict, 
                    option=orjson.OPT_INDENT_2 | orjson.OPT_NON_STR_KEYS
                )
                with output_file.open('wb') as f:
                    f.write(json_bytes)
                self.logger.debug("Used orjson for optimized serialization")
                
            except ImportError:
                with output_file.open('w', encoding='utf-8') as f:
                    json.dump(
                        result_dict,
                        f,
                        indent=2,
                        ensure_ascii=False,
                        separators=(',', ':')
                    )
                self.logger.debug("Used standard json (orjson not available)")
            
            self.logger.info(f"Analysis result saved: {output_file}")
            return output_file
            
        except Exception as e:
            self.logger.error(f"Failed to save analysis result: {e}")
            raise
    
    async def analyze_template_for_existing_job(self, xlsm_path: Path, existing_job_id: Union[int, str]) -> Dict[str, Any]:
        """Analyze template for an existing job.
        
        Args:
            xlsm_path: Path to XLSM template file
            existing_job_id: ID of existing job to add analysis to
            
        Returns:
            Dictionary containing analysis results
        """
        existing_job_dir = Path("production_output") / str(existing_job_id)
        if not existing_job_dir.exists():
            raise ValueError(f"Existing job directory not found: {existing_job_dir}")
        
        self.logger.info(f"🔗 Adding template analysis to existing job {existing_job_id}")
        return await self.analyze_template(xlsm_path, job_id=existing_job_id)
    
    def get_analysis_capabilities(self) -> Dict[str, Any]:
        """Get analyzer capabilities and configuration."""
        return {
            'supported_formats': ['.xlsm', '.xlsx'],
            'modular_architecture': {
                'worksheet_detector': 'WorksheetDetector',
                'header_detector': 'HeaderDetector', 
                'column_extractor': 'ColumnExtractor',
                'performance_monitor': 'PerformanceMonitor',
                'validation_utils': 'ValidationUtils'
            },
            'features': {
                'async_processing': True,
                'performance_monitoring': self.enable_monitoring,
                'requirement_status_extraction': True,
                'early_termination': True,
                'memory_optimization': True
            }
        }


# Backward compatibility alias
XlsmTemplateAnalyzer = OptimizedXlsmTemplateAnalyzer