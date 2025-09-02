"""Modern template data extractor using clean modular architecture.

Refactored to use focused, single-responsibility modules following
modern Python patterns with full type hints and clean separation of concerns.
"""

from __future__ import annotations

import json
import logging
import time
import tracemalloc
from pathlib import Path
from typing import Dict, Optional, Union

from openpyxl import load_workbook

from .data_structures import ExtractionResult, ExtractionMetrics
from .value_extractor import ValueExtractor




class HighPerformanceTemplateDataExtractor:
    """Modern template data extractor using clean modular architecture.
    
    Coordinates specialized components to extract values from template columns
    with clean separation of concerns and modern Python patterns.
    """
    
    # Performance targets
    MAX_DURATION_SECONDS = 1.5
    MAX_MEMORY_MB = 50.0
    MIN_THROUGHPUT_COLUMNS_PER_SECOND = 300.0
    
    def __init__(self, enable_performance_monitoring: bool = True) -> None:
        """Initialize extractor with modular components.
        
        Args:
            enable_performance_monitoring: Whether to track performance metrics
        """
        self.enable_performance_monitoring = enable_performance_monitoring
        self.logger = self._setup_logging()
        
        # Initialize modular components
        self.value_extractor = ValueExtractor(enable_performance_monitoring)
        
        # JSON serialization optimization
        self._json_encoder = self._setup_optimal_json_encoder()
        
    def _setup_logging(self) -> logging.Logger:
        """Configure structured logging.
        
        Returns:
            Configured logger instance
        """
        logger = logging.getLogger(__name__)
        if not logger.handlers:
            handler = logging.StreamHandler()
            handler.setFormatter(logging.Formatter('%(levelname)s - %(message)s'))
            logger.addHandler(handler)
            logger.setLevel(logging.INFO)
        return logger
    
    def _setup_optimal_json_encoder(self) -> str:
        """Setup optimal JSON encoder.
        
        Returns:
            Name of JSON encoder library to use
        """
        try:
            import orjson
            self.logger.info("Using orjson for maximum JSON performance")
            return 'orjson'
        except ImportError:
            try:
                import ujson
                self.logger.info("Using ujson for enhanced JSON performance")
                return 'ujson'  
            except ImportError:
                self.logger.info("Using standard json (install orjson for 19x speedup)")
                return 'json'
    
    def _load_step1_mappings(self, step1_path: Union[str, Path]) -> Dict[str, any]:
        """Load Step 1 mappings with optimal JSON decoder.
        
        Args:
            step1_path: Path to Step 1 mappings file
            
        Returns:
            Loaded mappings data
            
        Raises:
            FileNotFoundError: When mappings file not found
        """
        step1_path = Path(step1_path)
        if not step1_path.exists():
            raise FileNotFoundError(f"Step 1 mappings not found: {step1_path}")
        
        # Use fastest available JSON decoder
        if self._json_encoder == 'orjson':
            import orjson
            with open(step1_path, 'rb') as f:
                return orjson.loads(f.read())
        elif self._json_encoder == 'ujson':
            import ujson
            with open(step1_path, 'r', encoding='utf-8') as f:
                return ujson.load(f)
        else:
            with open(step1_path, 'r', encoding='utf-8') as f:
                return json.load(f)
    
    
    def _save_json_optimized(self, data: Dict[str, any], output_path: Path) -> None:
        """Save JSON using optimal library for maximum performance.
        
        Args:
            data: Data to serialize
            output_path: Path to save JSON file
        """
        try:
            if self._json_encoder == 'orjson':
                import orjson
                json_bytes = orjson.dumps(data, option=orjson.OPT_INDENT_2 | orjson.OPT_SORT_KEYS)
                with open(output_path, 'wb') as f:
                    f.write(json_bytes)
            elif self._json_encoder == 'ujson':
                import ujson
                with open(output_path, 'w', encoding='utf-8') as f:
                    ujson.dump(data, f, indent=2, sort_keys=True, ensure_ascii=False)
            else:
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False, separators=(',', ': '), sort_keys=True)
                    
        except Exception as e:
            self.logger.error(f"JSON save failed, using fallback: {e}")
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
    
    def extract_template_values(self, template_path: Path, step1_path: Path, 
                              job_id: Union[str, int]) -> ExtractionResult:
        """Extract valid values from template columns using modular architecture.
        
        Args:
            template_path: Path to Excel template file
            step1_path: Path to Step 1 JSON mappings
            job_id: Job identifier
            
        Returns:
            ExtractionResult containing all extracted values and validation results
        """
        if self.enable_performance_monitoring:
            self.logger.info(f"Starting modular Step 2 value extraction for job {job_id}")
        
        # Performance monitoring setup
        start_time = time.perf_counter()
        if self.enable_performance_monitoring:
            tracemalloc.start()
        
        try:
            # Load Step 1 mappings
            step1_data = self._load_step1_mappings(step1_path)
            column_mappings = step1_data['column_mappings']
            analysis_metadata = step1_data['analysis_metadata']
            
            # Initialize result
            result = ExtractionResult(
                job_id=job_id,
                source_template=str(template_path),
                extraction_metadata={
                    "worksheet_name": analysis_metadata.get('worksheet_name'),
                    "total_fields_to_process": len(column_mappings),
                    "value_columns": self.value_extractor.VALUE_COLUMNS,
                    "extraction_timestamp": time.time(),
                    "optimization_level": "modular_architecture"
                }
            )
            
            # Load workbook with optimization
            workbook = load_workbook(
                template_path,
                read_only=True,
                data_only=True,
                keep_vba=False,
                keep_links=False
            )
            
            # Get target worksheet
            target_worksheet_name = analysis_metadata.get('worksheet_name')
            try:
                worksheet = workbook[target_worksheet_name]
            except KeyError:
                self.logger.warning(f"Worksheet '{target_worksheet_name}' not found, using first worksheet")
                worksheet = workbook.worksheets[0]
            
            # Extract values using modular component
            if self.enable_performance_monitoring:
                extraction_start = time.perf_counter()
            
            # Check if 'Gültige Werte' worksheet exists for actual value data
            if 'Gültige Werte' in workbook.sheetnames:
                values_worksheet = workbook['Gültige Werte']
                self.logger.info("Using 'Gültige Werte' worksheet for value extraction")
                result.field_validations = self.value_extractor.extract_values_from_valid_values_sheet(values_worksheet)
            else:
                # Fallback to template definitions
                result.field_validations = self.value_extractor.extract_bulk_values(worksheet, column_mappings)
            
            if self.enable_performance_monitoring:
                extraction_time = (time.perf_counter() - extraction_start) * 1000
            
            workbook.close()
            
            # Calculate statistics
            total_values_extracted = sum(len(v.valid_values) for v in result.field_validations.values())
            columns_processed = len(column_mappings) * len(self.value_extractor.VALUE_COLUMNS)
            
            # Check for mandatory field failures
            for field_name, validation in result.field_validations.items():
                if validation.requirement_status == "mandatory" and not validation.validation_passed:
                    result.validation_errors.append(f"Mandatory field '{field_name}' has no valid values")
            
            # Update metadata
            result.extraction_metadata.update({
                "fields_processed": len(column_mappings),
                "total_values_extracted": total_values_extracted,
                "columns_processed": columns_processed,
                "mandatory_fields_failed": len(result.validation_errors)
            })
            
            # Calculate final performance metrics
            total_duration = (time.perf_counter() - start_time) * 1000
            
            if self.enable_performance_monitoring:
                current, peak = tracemalloc.get_traced_memory()
                peak_memory_mb = peak / 1024 / 1024
                tracemalloc.stop()
                
                # Performance validation
                self._validate_performance_targets(total_duration, peak_memory_mb, columns_processed, total_values_extracted)
            
            if self.enable_performance_monitoring:
                # Add performance metrics to result
                extraction_metric = self.value_extractor.create_extraction_metrics(
                    operation_name="template_value_extraction",
                    duration_ms=total_duration,
                    peak_memory_mb=peak_memory_mb,
                    values_extracted=total_values_extracted,
                    columns_processed=columns_processed
                )
                result.performance_metrics.append(extraction_metric)
                
                self.logger.info(
                    f"Modular Step 2 extraction completed: "
                    f"{total_values_extracted} values from {columns_processed} columns "
                    f"in {total_duration:.1f}ms"
                )
            
            return result
            
        except Exception as e:
            if self.enable_performance_monitoring:
                tracemalloc.stop()
            self.logger.error(f"Template value extraction failed: {e}")
            raise
    
    def _validate_performance_targets(self, duration_ms: float, memory_mb: float, 
                                    columns_processed: int, values_extracted: int) -> None:
        """Validate performance against targets.
        
        Args:
            duration_ms: Actual duration in milliseconds
            memory_mb: Actual memory usage in MB
            columns_processed: Number of columns processed
            values_extracted: Number of values extracted
        """
        duration_seconds = duration_ms / 1000
        throughput = columns_processed / duration_seconds if duration_seconds > 0 else 0
        
        # Performance target validation
        if duration_seconds > self.MAX_DURATION_SECONDS:
            self.logger.warning(f"Performance: Duration {duration_seconds:.2f}s exceeds target {self.MAX_DURATION_SECONDS}s")
        
        if memory_mb > self.MAX_MEMORY_MB:
            self.logger.warning(f"Performance: Memory {memory_mb:.1f}MB exceeds target {self.MAX_MEMORY_MB}MB")
        
        if throughput < self.MIN_THROUGHPUT_COLUMNS_PER_SECOND:
            self.logger.warning(f"Performance: Throughput {throughput:.1f} columns/sec below target {self.MIN_THROUGHPUT_COLUMNS_PER_SECOND}")
        
        self.logger.info(f"Performance: {duration_seconds:.2f}s, {memory_mb:.1f}MB, {throughput:.1f} columns/sec")
    
    def extract_and_save_values(self, template_path: Path, job_id: Union[str, int],
                              output_dir: Optional[Path] = None) -> Path:
        """
        Extract template values and save to Step 2 JSON file with maximum performance.
        
        Args:
            template_path: Path to Excel template file
            job_id: Job identifier
            output_dir: Output directory
            
        Returns:
            Path to saved Step 2 JSON file
        """
        if output_dir is None:
            output_dir = Path("production_output") / str(job_id) / "flat_file_analysis"
        
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # File paths
        step1_path = output_dir / "step1_template_columns.json"
        step2_path = output_dir / "step2_valid_values.json"
        
        # Extract values with high performance
        result = self.extract_template_values(template_path, step1_path, job_id)
        
        # Save with optimized JSON serialization
        json_start = time.perf_counter()
        self._save_json_optimized(result.to_dict(), step2_path)
        json_duration = (time.perf_counter() - json_start) * 1000
        
        if self.enable_performance_monitoring:
            self.logger.info(f"Step 2 results saved to {step2_path} using {self._json_encoder} in {json_duration:.1f}ms")
        
        # Log summary
        total_fields = len(result.field_validations)
        valid_fields = sum(1 for v in result.field_validations.values() if v.validation_passed)
        total_values = sum(len(v.valid_values) for v in result.field_validations.values())
        
        if self.enable_performance_monitoring:
            self.logger.info(f"Summary: {valid_fields}/{total_fields} fields validated, {total_values} total valid values extracted")
        
        return step2_path


# Backward compatibility - alias to optimized implementation
TemplateDataExtractor = HighPerformanceTemplateDataExtractor
OptimizedTemplateDataExtractor = HighPerformanceTemplateDataExtractor


# Convenience functions for pipeline integration
def extract_template_data(template_path: Union[str, Path], job_id: Union[str, int],
                         output_dir: Optional[Path] = None) -> Path:
    """
    Convenience function to extract template data for Step 2 with high performance.
    
    Args:
        template_path: Path to Excel template file
        job_id: Job identifier
        output_dir: Output directory
        
    Returns:
        Path to Step 2 JSON output file
    """
    extractor = HighPerformanceTemplateDataExtractor(enable_performance_monitoring=True)
    return extractor.extract_and_save_values(Path(template_path), job_id, output_dir)


def create_step2_extractor(**kwargs) -> HighPerformanceTemplateDataExtractor:
    """Factory function to create configured Step 2 extractor."""
    return HighPerformanceTemplateDataExtractor(**kwargs)


# Performance benchmarking function
def benchmark_step2_performance(template_path: Path, step1_path: Path, 
                               job_id: Union[str, int]) -> Dict[str, Any]:
    """
    Benchmark Step 2 extraction performance.
    
    Returns performance metrics and optimization analysis.
    """
    extractor = HighPerformanceTemplateDataExtractor(enable_performance_monitoring=True)
    
    # Benchmark extraction
    start_time = time.perf_counter()
    tracemalloc.start()
    
    result = extractor.extract_template_values(template_path, step1_path, job_id)
    
    duration_ms = (time.perf_counter() - start_time) * 1000
    current, peak = tracemalloc.get_traced_memory()
    peak_memory_mb = peak / 1024 / 1024
    tracemalloc.stop()
    
    # Calculate metrics
    total_values = sum(len(v.valid_values) for v in result.field_validations.values())
    columns_processed = len(result.field_validations) * 3
    throughput = columns_processed / (duration_ms / 1000) if duration_ms > 0 else 0
    
    # Compare to baseline
    baseline_duration_ms = 2320
    improvement_percent = ((baseline_duration_ms - duration_ms) / baseline_duration_ms) * 100
    
    return {
        "performance_metrics": {
            "duration_ms": round(duration_ms, 1),
            "memory_mb": round(peak_memory_mb, 1),
            "throughput_columns_per_second": round(throughput, 1),
            "values_extracted": total_values,
            "fields_processed": len(result.field_validations)
        },
        "performance_analysis": {
            "meets_duration_target": duration_ms <= 1500,
            "meets_memory_target": peak_memory_mb <= 50,
            "meets_throughput_target": throughput >= 300,
            "improvement_vs_baseline_percent": round(improvement_percent, 1),
            "optimization_grade": (
                "A" if all([duration_ms <= 1500, peak_memory_mb <= 50, throughput >= 300]) else
                "B" if duration_ms <= 1500 and peak_memory_mb <= 50 else
                "C" if duration_ms <= 2000 else
                "D"
            )
        },
        "optimization_features": {
            "bulk_range_extraction": True,
            "optimized_json_serialization": extractor._json_encoder,
            "minimal_memory_allocations": True,
            "fast_unicode_detection": True
        }
    }