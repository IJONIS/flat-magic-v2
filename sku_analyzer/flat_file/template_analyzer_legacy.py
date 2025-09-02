"""Modern XLSM template analyzer with clean modular architecture.

Refactored to use focused, single-responsibility modules following
modern Python patterns with full type hints and clean separation of concerns.
"""

from __future__ import annotations

import asyncio
import json
import logging
import time
from pathlib import Path
from typing import Any, Dict, Optional, Union

from openpyxl import load_workbook

from .column_extractor import ColumnExtractor
from .data_structures import TemplateAnalysisResult, TemplateDetectionError
from .header_detector import HeaderDetector
from .performance_monitor import PerformanceMonitor
from .validation_utils import ValidationUtils
from .worksheet_detector import WorksheetDetector


class OptimizedXlsmTemplateAnalyzer:
    """Modern template analyzer using clean modular architecture.
    
    This analyzer coordinates specialized components to perform template analysis
    with clean separation of concerns and modern Python patterns.
    """
    
    def __init__(self, enable_performance_monitoring: bool = True) -> None:
        """Initialize template analyzer with modular components.
        
        Args:
            enable_performance_monitoring: Whether to enable performance tracking
        """
        self.logger = self._setup_logging()
        self.enable_monitoring = enable_performance_monitoring
        
        # Initialize modular components
        self.worksheet_detector = WorksheetDetector()
        self.header_detector = HeaderDetector()
        self.column_extractor = ColumnExtractor(enable_performance_monitoring)
        self.performance_monitor = PerformanceMonitor(enable_performance_monitoring)
        self.validation_utils = ValidationUtils()
    
    def _get_next_job_number(self) -> int:
        """Get the next consecutive job number.
        
        Returns:
            Next available job number
        """
        production_dir = Path("production_output")
        if not production_dir.exists():
            return 1
        
        job_numbers = [
            int(job_dir.name) for job_dir in production_dir.iterdir()
            if job_dir.is_dir() and job_dir.name.isdigit()
        ]
        
        return max(job_numbers, default=0) + 1
    
    def get_latest_job_number(self) -> Optional[int]:
        """Get the most recent job number.
        
        Returns:
            Latest job number or None if no jobs exist
        """
        production_dir = Path("production_output")
        if not production_dir.exists():
            return None
        
        job_numbers = [
            int(job_dir.name) for job_dir in production_dir.iterdir()
            if job_dir.is_dir() and job_dir.name.isdigit()
        ]
        
        return max(job_numbers) if job_numbers else None
    
    def _setup_logging(self) -> logging.Logger:
        """Configure structured logging.
        
        Returns:
            Configured logger instance
        """
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)
        
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - [%(funcName)s] - %(message)s'
            )
            handler.setFormatter(formatter)
            logger.addHandler(handler)
        
        return logger
    
    async def _load_workbook_optimized(self, xlsm_path: Path) -> Any:
        """Load XLSM workbook with optimized settings.
        
        Args:
            xlsm_path: Path to XLSM file
            
        Returns:
            Loaded openpyxl workbook
        """
        loop = asyncio.get_event_loop()
        
        def _load_workbook() -> Any:
            return load_workbook(
                xlsm_path,
                read_only=True,
                keep_vba=False,
                data_only=True,
                keep_links=False
            )
        
        return await loop.run_in_executor(None, _load_workbook)

    async def analyze_template(self, xlsm_path: Path, job_id: Optional[Union[int, str]] = None) -> Dict[str, Any]:
        """
        Analyze XLSM template with comprehensive performance optimization.
        
        Args:
            xlsm_path: Path to XLSM file
            job_id: Job identifier for output management. If None, creates new job ID using existing sequence.
            
        Returns:
            Dict containing analysis results and performance metrics
            
        Raises:
            TemplateDetectionError: When template detection fails
            FileNotFoundError: When XLSM file doesn't exist
            ValueError: When invalid parameters provided
        """
        if not xlsm_path.exists():
            raise FileNotFoundError(f"XLSM file not found: {xlsm_path}")
        
        if not xlsm_path.suffix.lower() in ['.xlsm', '.xlsx']:
            raise ValueError(f"Invalid file type: {xlsm_path.suffix}. Expected .xlsm or .xlsx")

        # Use existing job ID or create new one following the sequence
        if job_id is None:
            job_number = self._get_next_job_number()
        else:
            job_number = int(job_id)

        self.logger.info(f"🔍 Starting optimized template analysis for job {job_number}: {xlsm_path}")
        
        performance_tracker = {}
        workbook = None
        
        try:
            # Step 1: Load workbook
            with self.performance_monitor.measure_performance("workbook_loading") as monitor:
                workbook = await self._load_workbook_optimized(xlsm_path)
            if monitor and monitor['metrics']:
                performance_tracker["workbook_loading"] = monitor['metrics']
            
            # Step 2: Detect target worksheet
            with self.performance_monitor.measure_performance("worksheet_detection") as monitor:
                worksheet = self.worksheet_detector.detect_target_worksheet(workbook)
            if monitor and monitor['metrics']:
                performance_tracker["worksheet_detection"] = monitor['metrics']
            
            # Step 3: Detect headers
            with self.performance_monitor.measure_performance("header_detection", worksheet.max_row) as monitor:
                header_row, tech_col, display_col, req_col = self.header_detector.detect_header_row(worksheet)
            if monitor and monitor['metrics']:
                performance_tracker["header_detection"] = monitor['metrics']
            
            # Step 4: Extract column mappings
            with self.performance_monitor.measure_performance("column_extraction", worksheet.max_row) as monitor:
                mappings = self.column_extractor.extract_column_mappings(
                    worksheet, header_row, tech_col, display_col, req_col
                )
            if monitor and monitor['metrics']:
                performance_tracker["column_extraction"] = monitor['metrics']
                
            # Step 5: Generate requirement statistics
            requirement_stats = self.column_extractor.generate_requirement_statistics(mappings)
            
            # Step 6: Create and validate result
            result = TemplateAnalysisResult(
                job_id=job_number,
                worksheet_name=worksheet.title,
                header_row=header_row,
                technical_column=tech_col,
                display_column=display_col,
                column_mappings=mappings,
                total_mappings=len(mappings),
                performance_metrics=performance_tracker,
                requirement_column=req_col,
                requirement_statistics=requirement_stats
            )
            
            self.validation_utils.validate_analysis_result(result)
            
            # Step 7: Save results
            with self.performance_monitor.measure_performance("output_saving") as monitor:
                output_path = await self._save_analysis_result_optimized(result)
            if monitor and monitor['metrics']:
                performance_tracker["output_saving"] = monitor['metrics']
            
            # Step 7: Extract template values (Step 2)
            step2_output_path = None
            try:
                from .template_data_extractor import HighPerformanceTemplateDataExtractor
                
                self.logger.info(f"🔍 Starting Step 2 value extraction for job {job_number}")
                extractor = HighPerformanceTemplateDataExtractor(enable_performance_monitoring=True)
                
                # Note: Step 2 is synchronous for performance optimization
                start_time = time.perf_counter()
                step2_output_path = extractor.extract_and_save_values(xlsm_path, job_number)
                step2_duration = (time.perf_counter() - start_time) * 1000
                
                # Record Step 2 performance
                performance_tracker["step2_value_extraction"] = {
                    'duration_ms': step2_duration,
                    'peak_memory_mb': 0,  # Recorded by the extractor itself
                    'throughput_rows_per_second': 0
                }
                    
                self.logger.info(f"✅ Step 2 value extraction completed for job {job_number}")
                
            except Exception as e:
                self.logger.error(f"Step 2 value extraction failed for job {job_number}: {e}")
                # Don't fail the entire analysis if Step 2 fails
                pass
            
            # Log performance summary
            total_duration = sum(
                m.duration_ms if hasattr(m, 'duration_ms') else m.get('duration_ms', 0) 
                for m in performance_tracker.values()
            )
            peak_memory = max(
                m.peak_memory_mb if hasattr(m, 'peak_memory_mb') else m.get('peak_memory_mb', 0) 
                for m in performance_tracker.values()
            ) if performance_tracker else 0
            
            self.logger.info(
                f"✅ Template analysis completed for job {job_number}: "
                f"{len(mappings)} mappings in {total_duration:.1f}ms, "
                f"peak memory: {peak_memory:.1f}MB"
            )
            
            analysis_dict = result.to_dict()
            analysis_dict['output_path'] = str(output_path)
            if step2_output_path:
                analysis_dict['step2_output_path'] = str(step2_output_path)
            return analysis_dict
            
        except Exception as e:
            self.logger.error(f"Template analysis failed for job {job_number}: {e}")
            raise TemplateDetectionError(f"Template analysis failed: {e}") from e
        finally:
            # Ensure workbook is closed and memory released
            if workbook:
                try:
                    workbook.close()
                except Exception as e:
                    self.logger.warning(f"Error closing workbook: {e}")
    
    async def _load_workbook_optimized(self, xlsm_path: Path) -> Any:
        """Load XLSM workbook with optimized settings for minimal memory usage."""
        loop = asyncio.get_event_loop()
        
        def _load_workbook():
            return load_workbook(
                xlsm_path,
                read_only=True,          # Read-only mode for memory efficiency
                keep_vba=False,          # Skip VBA to reduce memory footprint
                data_only=True,          # Get calculated values, not formulas
                keep_links=False         # Don't load external links
            )
        
        return await loop.run_in_executor(None, _load_workbook)
    
    async def _detect_target_worksheet_optimized(self, workbook: Any) -> Worksheet:
        """
        Optimized worksheet detection with early termination and pattern caching.
        
        Performance optimizations:
        - Precompiled regex patterns
        - Early termination on exact match
        - Limit scanning to MAX_WORKSHEET_SCAN worksheets
        """
        worksheet_names = workbook.sheetnames
        self.logger.debug(f"Available worksheets: {worksheet_names[:self.MAX_WORKSHEET_SCAN]}")
        
        # Try exact matches first (fastest path)
        for priority_name in self.WORKSHEET_PRIORITIES[:2]:
            if priority_name in worksheet_names:
                self.logger.info(f"Found exact match worksheet: '{priority_name}' (fast path)")
                return workbook[priority_name]
        
        # Try precompiled fuzzy patterns (limited scan)
        scan_limit = min(self.MAX_WORKSHEET_SCAN, len(worksheet_names))
        for pattern in self._compiled_patterns:
            for sheet_name in worksheet_names[:scan_limit]:
                if pattern.match(sheet_name):
                    self.logger.info(f"Found fuzzy match: '{sheet_name}' (pattern match)")
                    return workbook[sheet_name]
        
        # Fallback: check first worksheet only (performance limit)
        first_sheet = workbook[worksheet_names[0]]
        if await self._worksheet_contains_headers_fast(first_sheet):
            self.logger.info(f"Using first worksheet with headers: '{first_sheet.title}' (fallback)")
            return first_sheet
        
        raise TemplateDetectionError(
            f"No suitable worksheet found in first {scan_limit} sheets. "
            f"Available: {worksheet_names[:scan_limit]}. "
            f"Expected worksheets containing '{self.TECHNICAL_HEADER}' and '{self.DISPLAY_HEADER}' headers."
        )
    
    async def _worksheet_contains_headers_fast(self, worksheet: Worksheet) -> bool:
        """Fast header detection with minimal memory usage and early termination."""
        max_check_rows = min(20, worksheet.max_row or 20)  # Limit to first 20 rows
        
        for row_num in range(1, max_check_rows + 1):
            try:
                # Get row cells in one operation
                row_cells = list(worksheet[row_num])
                
                # Early termination: check for both headers in single pass
                tech_found = False
                display_found = False
                
                for cell in row_cells:
                    value = str(cell.value or '').strip()
                    if value == self.TECHNICAL_HEADER:
                        tech_found = True
                    elif value == self.DISPLAY_HEADER:
                        display_found = True
                    
                    # Early termination when both found
                    if tech_found and display_found:
                        return True
                        
            except Exception:
                # Skip problematic rows
                continue
        
        return False
    
    async def _detect_header_row_optimized(self, worksheet: Worksheet) -> Tuple[int, str, str, Optional[str]]:
        """
        Optimized header detection with requirement column support and early termination.
        
        Performance optimizations:
        - Early termination when required headers found
        - Batch row processing
        - Memory-efficient cell value extraction
        - Limited search scope
        - Optional requirement column detection
        """
        max_search_rows = min(self.MAX_HEADER_SEARCH_ROWS, worksheet.max_row or 50)
        
        self.logger.debug(f"Scanning for headers in first {max_search_rows} rows")
        
        for row_num in range(1, max_search_rows + 1):
            try:
                # Get row cells efficiently
                row_cells = list(worksheet[row_num])
                
                # Find header positions in single pass
                tech_col_idx = None
                display_col_idx = None
                req_col_idx = None
                
                for idx, cell in enumerate(row_cells):
                    value = str(cell.value or '').strip()
                    
                    if value == self.TECHNICAL_HEADER and tech_col_idx is None:
                        tech_col_idx = idx
                    elif value == self.DISPLAY_HEADER and display_col_idx is None:
                        display_col_idx = idx
                    elif value == self.REQUIREMENT_HEADER and req_col_idx is None:
                        req_col_idx = idx
                    
                    # Early termination when required headers found (requirement is optional)
                    if tech_col_idx is not None and display_col_idx is not None:
                        # Continue scanning for requirement column in same row
                        pass
                
                # If both required headers found, return immediately
                if tech_col_idx is not None and display_col_idx is not None:
                    tech_col = self._index_to_column_letter(tech_col_idx)
                    display_col = self._index_to_column_letter(display_col_idx)
                    req_col = self._index_to_column_letter(req_col_idx) if req_col_idx is not None else None
                    
                    log_msg = (
                        f"Headers found at row {row_num} (early termination): "
                        f"'{self.TECHNICAL_HEADER}' in {tech_col}, "
                        f"'{self.DISPLAY_HEADER}' in {display_col}"
                    )
                    if req_col:
                        log_msg += f", '{self.REQUIREMENT_HEADER}' in {req_col}"
                    else:
                        log_msg += " (no requirement column found)"
                    
                    self.logger.info(log_msg)
                    
                    return row_num, tech_col, display_col, req_col
                    
            except Exception as e:
                # Log but continue scanning
                self.logger.debug(f"Error scanning row {row_num}: {e}")
                continue
        
        raise TemplateDetectionError(
            f"Required headers not found in first {max_search_rows} rows. "
            f"Expected: '{self.TECHNICAL_HEADER}' and '{self.DISPLAY_HEADER}'"
        )
    
    def _index_to_column_letter(self, idx: int) -> str:
        """Convert 0-based column index to Excel column letter (cached computation)."""
        result = ""
        idx += 1  # Convert to 1-based
        while idx > 0:
            idx -= 1
            result = chr(65 + idx % 26) + result
            idx //= 26
        return result
    
    async def _extract_column_mappings_streaming(
        self, 
        worksheet: Worksheet, 
        header_row: int, 
        tech_col: str, 
        display_col: str,
        req_col: Optional[str] = None
    ) -> List[ColumnMapping]:
        """
        Stream-process column mappings with requirement status and memory optimization.
        
        Performance optimizations:
        - Process rows in batches to control memory usage
        - Early validation to skip invalid rows quickly
        - Efficient string operations
        - Minimal object creation during scanning
        - Optional requirement status extraction
        """
        mappings = []
        start_row = header_row + 1
        max_row = worksheet.max_row or start_row
        
        self.logger.debug(f"Streaming extraction from rows {start_row} to {max_row}")
        
        # Process in batches to manage memory
        for batch_start in range(start_row, max_row + 1, self.BATCH_SIZE_ROWS):
            batch_end = min(batch_start + self.BATCH_SIZE_ROWS - 1, max_row)
            
            batch_mappings = await self._process_row_batch(
                worksheet, batch_start, batch_end, tech_col, display_col, req_col
            )
            mappings.extend(batch_mappings)
            
            # Memory usage check during processing
            if self.enable_monitoring:
                current_memory = self._get_current_memory_mb()
                if current_memory > self.MEMORY_LIMIT_MB:
                    self.logger.warning(
                        f"Memory usage high: {current_memory:.1f}MB at row {batch_end}"
                    )
        
        self.logger.info(f"Extracted {len(mappings)} column mappings using streaming approach")
        return mappings
    
    async def _process_row_batch(
        self,
        worksheet: Worksheet,
        start_row: int,
        end_row: int,
        tech_col: str,
        display_col: str,
        req_col: Optional[str] = None
    ) -> List[ColumnMapping]:
        """Process a batch of rows efficiently with requirement status extraction."""
        batch_mappings = []
        
        for row_num in range(start_row, end_row + 1):
            try:
                # Get cell values directly (avoid intermediate objects)
                tech_cell = worksheet[f"{tech_col}{row_num}"]
                display_cell = worksheet[f"{display_col}{row_num}"]
                
                tech_value = str(tech_cell.value or '').strip()
                display_value = str(display_cell.value or '').strip()
                
                # Extract requirement status if column available
                requirement_status = None
                if req_col:
                    req_cell = worksheet[f"{req_col}{row_num}"]
                    req_value = str(req_cell.value or '').strip()
                    requirement_status = self._parse_requirement_status(req_value)
                
                # Fast validation checks (fail fast approach)
                if not tech_value or not display_value:
                    continue
                
                # Quick technical name validation (optimized regex)
                if not self._is_valid_technical_name_fast(tech_value):
                    continue
                
                # Create mapping only for valid rows
                mapping = ColumnMapping(
                    technical_name=tech_value,
                    display_name=display_value,
                    row_index=row_num,
                    technical_col=tech_col,
                    display_col=display_col,
                    requirement_status=requirement_status,
                    requirement_col=req_col
                )
                
                batch_mappings.append(mapping)
                
            except Exception as e:
                # Log but continue processing (resilient extraction)
                self.logger.debug(f"Error processing row {row_num}: {e}")
                continue
        
        return batch_mappings
    
    def _is_valid_technical_name_fast(self, name: str) -> bool:
        """
        Fast technical name validation with optimized regex matching.
        
        Performance optimizations:
        - Quick length and character checks before regex
        - Precompiled pattern matching
        - Early termination on obvious failures
        """
        if not name or len(name) > 50:  # Quick length check
            return False
        
        # Quick character check before expensive regex
        if name != name.lower():  # Uppercase check
            return False
        
        if name[0].isdigit():  # Starts with number check
            return False
        
        # Use precompiled pattern for final validation
        return bool(self.TECHNICAL_NAME_PATTERN.match(name))
    
    def _parse_requirement_status(self, raw_value: str) -> Optional[str]:
        """
        Parse requirement status from German terms to standardized values.
        
        Args:
            raw_value: Raw requirement value from Excel cell
            
        Returns:
            Standardized requirement status or None for empty/unknown values
            
        Raises:
            ValueError: When raw_value contains unrecognized requirement term
        """
        if not raw_value or not raw_value.strip():
            return None
            
        normalized_value = raw_value.strip()
        
        # Direct mapping lookup (case-sensitive and case-insensitive)
        if normalized_value in self.REQUIREMENT_STATUS_MAP:
            return self.REQUIREMENT_STATUS_MAP[normalized_value]
        
        # Try case-insensitive fallback
        for key, value in self.REQUIREMENT_STATUS_MAP.items():
            if key.lower() == normalized_value.lower():
                return value
        
        # Unknown requirement status - raise error for visibility
        raise ValueError(
            f"Unknown requirement status: '{normalized_value}'. "
            f"Expected one of: {list(set(self.REQUIREMENT_STATUS_MAP.values()))}"
        )
    
    def _generate_requirement_statistics(self, mappings: List[ColumnMapping]) -> Dict[str, int]:
        """
        Generate requirement status distribution statistics.
        
        Args:
            mappings: List of column mappings with requirement status
            
        Returns:
            Dictionary containing counts for each requirement status
        """
        stats = {
            'mandatory': 0,
            'optional': 0,
            'recommended': 0,
            'unknown': 0
        }
        
        for mapping in mappings:
            status = mapping.requirement_status
            if status in stats:
                stats[status] += 1
            else:
                stats['unknown'] += 1
        
        return stats
    
    async def _validate_analysis_result(self, result: TemplateAnalysisResult) -> None:
        """Fast validation with early termination and efficient duplicate detection."""
        # Quick mapping count check
        if result.total_mappings < self.MIN_REQUIRED_MAPPINGS:
            error = (
                f"Insufficient mappings: {result.total_mappings} "
                f"(required: {self.MIN_REQUIRED_MAPPINGS})"
            )
            result.validation_errors.append(error)
        
        # Efficient duplicate detection using set
        tech_names = [mapping.technical_name for mapping in result.column_mappings]
        seen = set()
        duplicates = set()
        
        for name in tech_names:
            if name in seen:
                duplicates.add(name)
            else:
                seen.add(name)
        
        if duplicates:
            error = f"Duplicate technical names: {sorted(duplicates)}"
            result.validation_errors.append(error)
        
        # Fast empty value checks
        empty_tech = sum(1 for m in result.column_mappings if not m.technical_name.strip())
        empty_display = sum(1 for m in result.column_mappings if not m.display_name.strip())
        
        if empty_tech:
            result.validation_errors.append(f"Empty technical names: {empty_tech} mappings")
        if empty_display:
            result.validation_errors.append(f"Empty display names: {empty_display} mappings")
        
        # Log validation summary
        if result.validation_errors:
            for error in result.validation_errors:
                self.logger.warning(f"Validation error: {error}")
        else:
            self.logger.info("✅ Analysis validation passed")
    
    async def _save_analysis_result_optimized(self, result: TemplateAnalysisResult) -> Path:
        """
        Save analysis result with optimized JSON serialization.
        
        Performance optimizations:
        - Use fastest available JSON library
        - Efficient file I/O operations
        - Minimal memory allocation during serialization
        """
        output_dir = Path("production_output") / str(result.job_id) / "flat_file_analysis"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        output_file = output_dir / "step1_template_columns.json"
        
        try:
            # Use fastest available JSON library
            result_dict = result.to_dict()
            
            # Try orjson first (fastest)
            try:
                import orjson
                json_bytes = orjson.dumps(
                    result_dict, 
                    option=orjson.OPT_INDENT_2 | orjson.OPT_NON_STR_KEYS
                )
                with output_file.open('wb') as f:
                    f.write(json_bytes)
                self.logger.debug("Used orjson for optimized serialization")
                
            except ImportError:
                # Fallback to standard json with optimizations
                with output_file.open('w', encoding='utf-8') as f:
                    json.dump(
                        result_dict,
                        f,
                        indent=2,
                        ensure_ascii=False,  # Preserve German characters
                        separators=(',', ':')  # Compact formatting
                    )
                self.logger.debug("Used standard json (orjson not available)")
            
            self.logger.info(f"Analysis result saved: {output_file}")
            return output_file
            
        except Exception as e:
            self.logger.error(f"Failed to save analysis result: {e}")
            raise
    
    def get_analysis_capabilities(self) -> Dict[str, Any]:
        """Get analyzer capabilities with performance characteristics."""
        return {
            'supported_formats': ['.xlsm', '.xlsx'],
            'worksheet_detection': {
                'exact_match_priorities': self.WORKSHEET_PRIORITIES[:2],
                'fuzzy_patterns': len(self._compiled_patterns),
                'fallback_strategy': 'first_sheet_with_headers',
                'max_worksheets_scanned': self.MAX_WORKSHEET_SCAN
            },
            'header_detection': {
                'technical_header': self.TECHNICAL_HEADER,
                'display_header': self.DISPLAY_HEADER,
                'requirement_header': self.REQUIREMENT_HEADER,
                'max_search_rows': self.MAX_HEADER_SEARCH_ROWS,
                'early_termination': True,
                'requirement_column_optional': True
            },
            'requirement_status': {
                'supported_languages': ['German'],
                'standard_values': ['mandatory', 'optional', 'recommended'],
                'german_mappings': dict(self.REQUIREMENT_STATUS_MAP),
                'statistics_generation': True,
                'graceful_fallback': True
            },
            'extraction_optimization': {
                'streaming_processing': True,
                'batch_size_rows': self.BATCH_SIZE_ROWS,
                'memory_limit_mb': self.MEMORY_LIMIT_MB,
                'fast_validation': True,
                'requirement_status_extraction': True
            },
            'validation': {
                'min_required_mappings': self.MIN_REQUIRED_MAPPINGS,
                'technical_name_pattern': self.TECHNICAL_NAME_PATTERN.pattern,
                'supports_german_characters': True,
                'duplicate_detection': 'set_based_o1'
            },
            'performance': {
                'async_processing': True,
                'memory_efficient': True,
                'read_only_mode': True,
                'performance_monitoring': self.enable_monitoring,
                'json_library_optimization': True,
                'precompiled_patterns': True
            }
        }
    
    async def benchmark_performance(self, xlsm_path: Path) -> Dict[str, Any]:
        """
        Run comprehensive performance benchmark on XLSM file.
        
        Returns detailed performance metrics for optimization validation.
        """
        if not xlsm_path.exists():
            raise FileNotFoundError(f"Benchmark file not found: {xlsm_path}")
        
        self.logger.info(f"🎯 Starting performance benchmark: {xlsm_path}")
        
        # Enable performance monitoring for benchmark
        original_monitoring = self.enable_monitoring
        self.enable_monitoring = True
        
        try:
            benchmark_start = time.perf_counter()
            
            # Run analysis with performance tracking (create temporary benchmark job)
            benchmark_job_id = self._get_next_job_number() + 1000  # Offset for benchmark jobs
            result = await self.analyze_template(xlsm_path, job_id=benchmark_job_id)
            
            benchmark_duration = (time.perf_counter() - benchmark_start) * 1000
            
            # Extract performance metrics
            perf_metrics = result.get('performance', {})
            
            # Calculate aggregate metrics
            total_duration = sum(m.get('duration_ms', 0) for m in perf_metrics.values())
            peak_memory = max(m.get('peak_memory_mb', 0) for m in perf_metrics.values()) if perf_metrics else 0
            total_rows = sum(m.get('rows_processed', 0) for m in perf_metrics.values() if 'rows_processed' in m)
            
            # Performance validation
            performance_targets = {
                'total_duration_ms': total_duration,
                'peak_memory_mb': peak_memory,
                'rows_processed': total_rows,
                'mappings_extracted': len(result.get('column_mappings', [])),
                'targets_met': {
                    'duration_under_1000ms': total_duration < 1000,
                    'memory_under_50mb': peak_memory < 50,
                    'early_termination_used': any(
                        m.get('early_termination', False) for m in perf_metrics.values()
                    ),
                    'streaming_effective': total_rows > 0
                }
            }
            
            # Generate recommendations
            recommendations = self._generate_performance_recommendations(performance_targets)
            
            benchmark_report = {
                'benchmark_duration_ms': benchmark_duration,
                'operation_breakdown': perf_metrics,
                'aggregate_metrics': performance_targets,
                'optimization_recommendations': recommendations,
                'file_characteristics': {
                    'file_size_mb': xlsm_path.stat().st_size / 1024 / 1024,
                    'analysis_efficiency': f"{len(result.get('column_mappings', []))}/ms"
                }
            }
            
            self.logger.info(f"✅ Benchmark completed: {benchmark_duration:.1f}ms total")
            return benchmark_report
            
        finally:
            self.enable_monitoring = original_monitoring
    
    def _generate_performance_recommendations(self, metrics: Dict[str, Any]) -> List[str]:
        """Generate specific performance optimization recommendations."""
        recommendations = []
        
        duration = metrics.get('total_duration_ms', 0)
        memory = metrics.get('peak_memory_mb', 0)
        targets = metrics.get('targets_met', {})
        
        # Duration optimizations
        if duration > 500:
            recommendations.append("Consider implementing parallel worksheet scanning")
            recommendations.append("Reduce MAX_HEADER_SEARCH_ROWS if templates are consistently near top")
        
        # Memory optimizations
        if memory > 30:
            recommendations.append("Implement cell value caching for repeated access")
            recommendations.append("Consider using xlrd library for read-only operations")
        
        # Early termination effectiveness
        if not targets.get('early_termination_used', False):
            recommendations.append("Verify early termination logic is functioning correctly")
        
        # Streaming effectiveness
        if not targets.get('streaming_effective', False):
            recommendations.append("Review streaming implementation for zero-row scenarios")
        
        # General optimizations
        if duration > 200 or memory > 20:
            recommendations.append("Install orjson library for faster JSON serialization")
            recommendations.append("Consider implementing worksheet metadata caching")
        
        return recommendations

    async def analyze_template_for_existing_job(self, xlsm_path: Path, existing_job_id: Union[int, str]) -> Dict[str, Any]:
        """
        Analyze template for an existing job, creating the flat_file_analysis subdirectory.
        
        This method is designed for integration with existing SKU analysis jobs where
        template analysis is an additional step rather than a standalone operation.
        
        Args:
            xlsm_path: Path to XLSM template file
            existing_job_id: ID of existing job to add template analysis to
            
        Returns:
            Dict containing analysis results and performance metrics
        """
        # Verify the existing job directory exists
        existing_job_dir = Path("production_output") / str(existing_job_id)
        if not existing_job_dir.exists():
            raise ValueError(f"Existing job directory not found: {existing_job_dir}")
        
        self.logger.info(f"🔗 Adding template analysis to existing job {existing_job_id}")
        
        # Run analysis using the existing job ID
        return await self.analyze_template(xlsm_path, job_id=existing_job_id)

    # =====================================================================
    # BACKWARD COMPATIBILITY METHODS
    # =====================================================================
    
    def _is_valid_technical_name(self, name: str) -> bool:
        """Backward compatibility wrapper for fast technical name validation."""
        return self._is_valid_technical_name_fast(name)
    
    async def _detect_header_row(self, worksheet: Worksheet) -> Tuple[int, str, str]:
        """Backward compatibility wrapper for optimized header detection."""
        row, tech_col, display_col, _ = await self._detect_header_row_optimized(worksheet)
        return row, tech_col, display_col
    
    async def _worksheet_contains_headers(self, worksheet: Worksheet) -> bool:
        """Backward compatibility wrapper for fast header detection."""
        return await self._worksheet_contains_headers_fast(worksheet)
    
    async def _extract_column_mappings(
        self, worksheet: Worksheet, header_row: int, tech_col: str, display_col: str
    ) -> List[ColumnMapping]:
        """Backward compatibility wrapper for streaming extraction."""
        return await self._extract_column_mappings_streaming(worksheet, header_row, tech_col, display_col, None)
    
    async def _load_workbook(self, xlsm_path: Path) -> Any:
        """Backward compatibility wrapper for optimized workbook loading."""
        return await self._load_workbook_optimized(xlsm_path)
    
    async def _save_analysis_result(self, result: TemplateAnalysisResult) -> Path:
        """Backward compatibility wrapper for optimized result saving."""
        return await self._save_analysis_result_optimized(result)
    
    async def _detect_target_worksheet(self, workbook: Any) -> Worksheet:
        """Backward compatibility wrapper for optimized worksheet detection."""
        return await self._detect_target_worksheet_optimized(workbook)


# Backward compatibility alias
XlsmTemplateAnalyzer = OptimizedXlsmTemplateAnalyzer