#!/usr/bin/env python3
"""Test script to verify dynamic column detection is working correctly."""

import logging
from pathlib import Path
from openpyxl import load_workbook

from sku_analyzer.flat_file.value_extractor import ValueExtractor

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def test_column_detection():
    """Test the dynamic column detection for Land/Region der Herkunft field."""
    
    # Find the template file
    template_path = None
    for possible_path in [
        "EIKO-Datenimport-Vorlage_modifiziert.xlsm",
        "template/EIKO-Datenimport-Vorlage_modifiziert.xlsm",
        "input/EIKO-Datenimport-Vorlage_modifiziert.xlsm",
        "test-files/PANTS (3).xlsm"
    ]:
        full_path = Path(possible_path)
        if full_path.exists():
            template_path = full_path
            break
    
    if not template_path:
        logger.error("Template file not found")
        return
    
    logger.info(f"Using template: {template_path}")
    
    # Load workbook and get Gültige Werte sheet
    workbook = load_workbook(template_path, data_only=True)
    if "Gültige Werte" not in workbook.sheetnames:
        logger.error("'Gültige Werte' sheet not found")
        return
    
    worksheet = workbook["Gültige Werte"]
    logger.info(f"Worksheet max_column: {worksheet.max_column} (column {worksheet.cell(1, worksheet.max_column).coordinate[:-1]})")
    logger.info(f"Worksheet max_row: {worksheet.max_row}")
    
    # Find the Land/Region der Herkunft row
    land_region_row = None
    for row in range(2, min(worksheet.max_row + 1, 200)):
        field_name = worksheet[f'B{row}'].value
        if field_name and "Land/Region der Herkunft" in str(field_name):
            land_region_row = row
            break
    
    if not land_region_row:
        logger.error("Land/Region der Herkunft field not found")
        return
    
    logger.info(f"Found Land/Region der Herkunft at row {land_region_row}")
    
    # Test the current value extractor
    extractor = ValueExtractor()
    field_validations = extractor.extract_values_from_valid_values_sheet(worksheet)
    
    land_field = field_validations.get("Land/Region der Herkunft")
    if land_field:
        logger.info(f"Current extraction found {len(land_field.valid_values)} countries")
        logger.info(f"First 10 countries: {land_field.valid_values[:10]}")
        logger.info(f"Last 10 countries: {land_field.valid_values[-10:]}")
    
    # Manual scan to find all countries in the Land/Region row
    logger.info("\n=== Manual scan of Land/Region row ===")
    countries_found = []
    col_num = 3  # Start from column C
    consecutive_empty = 0
    max_col_with_data = 3
    
    while col_num <= worksheet.max_column:
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_num)
        cell_value = worksheet[f'{col_letter}{land_region_row}'].value
        
        if cell_value and str(cell_value).strip():
            countries_found.append(str(cell_value).strip())
            consecutive_empty = 0
            max_col_with_data = col_num
            logger.info(f"  Column {col_letter} ({col_num}): {str(cell_value).strip()}")
        else:
            consecutive_empty += 1
            # Don't stop early in this test - scan the entire row
        
        col_num += 1
    
    logger.info(f"\nManual scan results:")
    logger.info(f"  Total countries found: {len(countries_found)}")
    logger.info(f"  Maximum column with data: {get_column_letter(max_col_with_data)} ({max_col_with_data})")
    logger.info(f"  First country: {countries_found[0] if countries_found else 'None'}")
    logger.info(f"  Last country: {countries_found[-1] if countries_found else 'None'}")
    
    # Check if there's a difference between extraction and manual scan
    extracted_count = len(land_field.valid_values) if land_field else 0
    manual_count = len(countries_found)
    
    if extracted_count != manual_count:
        logger.warning(f"MISMATCH: Extractor found {extracted_count} countries, manual scan found {manual_count}")
        # Find missing countries
        extracted_set = set(land_field.valid_values) if land_field else set()
        manual_set = set(countries_found)
        missing = manual_set - extracted_set
        extra = extracted_set - manual_set
        
        if missing:
            logger.warning(f"  Missing from extraction: {list(missing)[:10]}...")
        if extra:
            logger.warning(f"  Extra in extraction: {list(extra)[:10]}...")
    else:
        logger.info("✅ Extraction matches manual scan!")

if __name__ == "__main__":
    test_column_detection()